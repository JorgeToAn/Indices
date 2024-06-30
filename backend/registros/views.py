from rest_framework.permissions import IsAuthenticated, IsAdminUser
from backend.permissions import IsAdminUserOrReadOnly

from rest_framework import generics, views
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes

from .serializers import IngresoSerializer, EgresoSerializer, TitulacionSerializer, LiberacionInglesSerializer
from .models import Ingreso, Egreso, Titulacion, LiberacionIngles
from .periodos import getPeriodoActual

from personal.models import Personal, obtenerFechaNac, obtenerGenero
from alumnos.models import Alumno
from carreras.models import Carrera
from planes.models import Plan

import openpyxl
import re

### INGRESO
class IngresoList(generics.ListCreateAPIView):
    queryset = Ingreso.objects.all()
    serializer_class = IngresoSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

class IngresoDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Ingreso.objects.all()
    serializer_class = IngresoSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

# FORMATO DE EXCEL [CURP, NO_CONTROL, PATERNO, MATERNO, NOMBRE, CARRERA, (PERIODO+TIPO)]
class IngresoUpload(views.APIView):
    parser_classes = [FileUploadParser]
    permission_classes = [IsAuthenticated&IsAdminUser]

    def to_dict(self, row):
        # regresa None si el renglon son solo celdas vacias
        for cell in row:
            if cell.value is not None:
                break
            return None

        if row[0].value is None:
            raise Exception('Se necesita un CURP')
        if row[1].value is None:
            raise Exception('Se necesita un no. de control')
        if row[4].value is None:
            raise Exception('Se necesita un nombre')
        if row[5].value is None:
            raise Exception('Se necesita una carrera')
        if row[6].value is None:
            raise Exception('Se necesita un tipo de ingreso')

        data = {
            'curp': str(row[0].value).strip(),
            'no_control': str(row[1].value).strip(),
            'paterno': str(row[2].value).strip(),
            'materno': str(row[3].value).strip(),
            'nombre': str(row[4].value).strip(),
            'carrera': str(row[5].value).strip(),
            'tipo': str(row[6].value).strip()[0:2],
        }
        return data

    def post(self, request, filename, format=None):
        ESTRUCTURA = [(r'^curp$', 'CURP'), (r'^no_control$', 'NO_CONTROL'), (r'^paterno$', 'PATERNO'), (r'^materno$', 'MATERNO'), (r'^nombre$', 'NOMBRE'), (r'^carrera$', 'CARRERA'), (r'^[12][0-9]{3}[13]$', 'NUMERO DE PERIODO')]
        file_obj = request.data['file']
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active

        results = {"errors": [], "created": 0}
        header_row = ws['A1':'G1'][0] # ws['A1':'G1'] regresa una tupla de renglones, pero solo necesitamos la primera

        # VALIDAR ESTRUCTURA DEL ARCHIVO COMO:
        # curp | no_control | paterno | materno | nombre | carrera | periodo
        for i, expresion in enumerate(ESTRUCTURA):
            match = re.match(expresion[0], str(header_row[i].value).lower())
            if match is None:
                return Response(status=400, data={'message': f'Se esperaba el campo {expresion[1]} pero se obtuvo {header_row[i].value}'})

        for row in ws.iter_rows(min_row=2):
            try:
                # se debe verificar que todos los campos tengan datos
                data = self.to_dict(row)
                if data is None:
                    continue
                # VALIDAR DATOS
                Personal.validate_curp(data['curp'])
                Alumno.validate_nocontrol(data['no_control'])
                carrera = Carrera.objects.get(pk=data['carrera'])

                personal, created_personal = Personal.objects.get_or_create(
                    curp=data['curp'],
                    paterno=data['paterno'],
                    materno=data['materno'],
                    nombre=data['nombre'],
                    fecha_nacimiento=obtenerFechaNac(data['curp']),
                    genero=obtenerGenero(data['curp'])
                )
                alumno = None
                try:
                    alumno = Alumno.objects.get(no_control=data['no_control'])
                    if alumno.plan.carrera.clave != data['carrera']:
                        results['errors'].append({'type': 'Carrera', 'message': f"La carrera {data['carrera']} no concuerda con el plan registrado {alumno.plan.clave}", 'row_index': row[0].row})
                        continue
                except Alumno.DoesNotExist:
                    plan = Plan.objects.filter(carrera=carrera).last()
                    alumno = Alumno.objects.create(
                        no_control=data['no_control'],
                        curp=personal,
                        plan=plan
                    )

                # CREAR REGISTRO DE INGRESO
                if not Ingreso.objects.filter(alumno=alumno, periodo=str(header_row[6].value), tipo=data['tipo']).exists():
                    ingreso = Ingreso(alumno=alumno, periodo=str(header_row[6].value), tipo=data['tipo'])
                    try:
                        ingreso.calcular_num_semestre()
                        ingreso.full_clean()
                        ingreso.save()
                        results['created'] += 1
                    except Exception as ex:
                        if ex.message_dict:
                            for err in ex.message_dict:
                                results['errors'].append({'type': err, 'message': ex.message_dict[err], 'row_index': row[0].row})
                        else:
                            results['errors'].append({'type': str(type(ex)), 'message': str(ex), 'row_index': row[0].row})
                        continue

            except Carrera.DoesNotExist as ex:
                results['errors'].append({'type': str(type(ex)), 'message': 'La carrera indicada no existe', 'row_index': row[0].row})
                continue
            except Exception as ex:
                results['errors'].append({'type': str(type(ex)), 'message': str(ex), 'row_index': row[0].row})
                continue
        return Response(status=200, data=results)

### EGRESO
class EgresoList(generics.ListCreateAPIView):
    queryset = Egreso.objects.all()
    serializer_class = EgresoSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

class EgresoDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Egreso.objects.all()
    serializer_class = EgresoSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

# FORMATO DE EXCEL [NO_CONTROL, PERIODO]
class EgresoUpload(views.APIView):
    parser_classes = [FileUploadParser]
    permission_classes = [IsAuthenticated&IsAdminUser]

    def to_dict(self, row):
        # regresa None si el renglon son solo celdas vacias
        for cell in row:
            if cell.value is not None:
                break
            return None

        if row[0].value is None:
            raise Exception('Se necesita un no. de control')

        data = {
            'no_control': str(row[0].value).strip(),
        }
        return data

    def post(self, request, filename, format=None):
        ESTRUCTURA = [(r'^no_control$', 'NO_CONTROL'), (r'^[12][0-9]{3}[13]$', 'PERIODO')]
        file_obj = request.data['file']
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active

        results = {"errors": [], "created": 0}
        header_row = ws['A1':'B1'][0] # ws['A1':'B1'] regresa una tupla de renglones, pero solo necesitamos la primera

        # VALIDAR ESTRUCTURA DEL ARCHIVO COMO:
        # no_control | periodo
        for i, expresion in enumerate(ESTRUCTURA):
            match = re.match(expresion[0], str(header_row[i].value).lower())
            if match is None:
                return Response(status=400, data={'message': f'Se esperaba el campo {expresion[i]} pero se obtuvo {header_row[i].value}'})

        for row in ws.iter_rows(min_row=2):
            try:
                data = self.to_dict(row)
                if data is None:
                    continue
                alumno = Alumno.objects.get(pk=data['no_control'])
                egresado = Egreso.objects.create(periodo=str(header_row[1].value), alumno=alumno)
                egresado.save()
                results['created'] += 1
            except Alumno.DoesNotExist as ex:
                results['errors'].append({'type': str(type(ex)), 'message': f'No se encontro un alumno con no. de control {data["no_control"]}', 'row_index': row[0].row})
            except Exception as ex:
                results['errors'].append({'type': str(type(ex)), 'message': str(ex), 'row_index': row[0].row})
        return Response(status=200, data=results)

### TITULACION
class TitulacionList(generics.ListCreateAPIView):
    queryset = Titulacion.objects.all()
    serializer_class = TitulacionSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

class TitulacionDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Titulacion.objects.all()
    serializer_class = TitulacionSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

# FORMATO DE EXCEL [NO_CONTROL, (PERIODO+TIPO)]
class TitulacionUpload(views.APIView):
    parser_classes = [FileUploadParser]
    permission_classes = [IsAuthenticated&IsAdminUser]

    def to_dict(self, row):
        # regresa None si el renglon son solo celdas vacias
        for cell in row:
            if cell.value is not None:
                break
            return None

        if row[0].value is None:
            raise Exception('Se necesita un no. de control')
        if row[1].value is None:
            raise Exception('Se necesita el tipo de titulación')

        data = {
            'no_control': str(row[0].value).strip(),
            'tipo_titulacion': str(row[1].value).strip()[0:2],
        }
        return data

    def post(self, request, filename, format=None):
        ESTRUCTURA = [(r'^no_control$', 'NO_CONTROL'), (r'^[12][0-9]{3}[13]$', 'NUMERO DE PERIODO')]
        file_obj = request.data['file']
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active

        results = {"errors": [], "created": 0}
        header_row = ws['A1':'B1'][0] # ws['A1':'B1'] regresa una tupla de renglones, pero solo necesitamos la primera

        # VALIDAR ESTRUCTURA DEL ARCHIVO COMO:
        # no_control | periodo
        for i, expresion in enumerate(ESTRUCTURA):
            match = re.match(expresion[0], str(header_row[i].value).lower())
            if match is None:
                return Response(status=400, data={'message': f'Se esperaba el campo {expresion[i]} pero se obtuvo {header_row[i].value}'})

        for row in ws.iter_rows(min_row=2):
            try:
                data = self.to_dict(row)
                if data is None:
                    continue
                alumno = Alumno.objects.get(pk=data['no_control'])
                titulacion, created = Titulacion.objects.get_or_create(periodo=str(header_row[1].value), tipo=data['tipo_titulacion'], alumno=alumno)
                if created:
                    results['created'] += 1
            except Alumno.DoesNotExist as ex:
                results['errors'].append({'type': str(type(ex)), 'message': f'No se encontro un alumno con no. de control {data["no_control"]}', 'row_index': row[0].row})
            except Exception as ex:
                results['errors'].append({'type': str(type(ex)), 'message': str(ex), 'row_index': row[0].row})
        return Response(status=200, data=results)

### LIBERACION DE INGLES
class LiberacionInglesList(generics.ListCreateAPIView):
    queryset = LiberacionIngles.objects.all()
    serializer_class = LiberacionInglesSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

class LiberacionInglesDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = LiberacionIngles.objects.all()
    serializer_class = LiberacionInglesSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

# FORMATO DE EXCEL [NO_CONTROL, PERIODO]
class LiberacionInglesUpload(views.APIView):
    parser_classes = [FileUploadParser]
    permission_classes = [IsAuthenticated&IsAdminUser]

    def to_dict(self, row):
        # regresa None si el renglon son solo celdas vacias
        for cell in row:
            if cell.value is not None:
                break
            return None

        if row[0].value is None:
            raise Exception('Se necesita un no. de control')

        data = {
            'no_control': str(row[0].value).strip(),
        }
        return data

    def post(self, request, filename, format=None):
        ESTRUCTURA = [(r'^no_control$', 'NO_CONTROL'), (r'^[12][0-9]{3}[13]$', 'NUMERO DE PERIODO')]
        file_obj = request.data['file']
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active

        results = {"errors": [], "created": 0}
        header_row = ws['A1':'B1'][0] # ws['A1':'B1'] regresa una tupla de renglones, pero solo necesitamos la primera

        # VALIDAR ESTRUCTURA DEL ARCHIVO COMO:
        # no_control | periodo
        for i, expresion in enumerate(ESTRUCTURA):
            match = re.match(expresion[0], str(header_row[i].value).lower())
            if match is None:
                return Response(status=400, data={'message': f'Se esperaba el campo {expresion[i]} pero se obtuvo {header_row[i].value}'})

        for row in ws.iter_rows(min_row=2):
            try:
                data = self.to_dict(row)
                if data is None:
                    continue
                alumno = Alumno.objects.get(pk=data['no_control'])
                liberacion, created = LiberacionIngles.objects.get_or_create(periodo=str(header_row[1].value), alumno=alumno)
                if created:
                    results['created'] += 1
            except Alumno.DoesNotExist as ex:
                results['errors'].append({'type': str(type(ex)), 'message': f'No se encontro un alumno con no. de control {data["no_control"]}', 'row_index': row[0].row})
            except Exception as ex:
                results['errors'].append({'type': str(type(ex)), 'message': str(ex), 'row_index': row[0].row})
        return Response(status=200, data=results)

### CORTE
@api_view(['POST',])
@permission_classes([IsAuthenticated&IsAdminUser])
def corte(request):
    periodo = getPeriodoActual()
    if not Ingreso.objects.contiene_corte(periodo) and not Egreso.objects.contiene_corte(periodo) and not Titulacion.objects.contiene_corte(periodo) and not LiberacionIngles.objects.contiene_corte(periodo):
        ingresos = Ingreso.objects.realizar_corte(periodo)
        egresos = Egreso.objects.realizar_corte(periodo)
        titulaciones = Titulacion.objects.realizar_corte(periodo)
        liberaciones = LiberacionIngles.objects.realizar_corte(periodo)
        return Response(status=200, data={'periodo': periodo, 'updated': {'ingresos': ingresos, 'egresos': egresos, 'titulaciones': titulaciones, 'liberaciones-ingles': liberaciones}})
    else:
        return Response(status=400, data={'periodo': periodo ,'message': f'No se puede realizar un corte ya que existen registros que pertenecen a un corte para el periodo {periodo}.'})

# Lee un renglon de excel y lo convierte a un diccionario de acuerdo a los campos del header_row
def row_to_dict(header_row, data_row):
    clean_header = clean_row(header_row)
    clean_data = clean_row(data_row)
    keywords = ['curp', 'no_control', 'paterno', 'materno', 'nombre', 'carrera']
    row_dict = {'periodos': []}
    for cell in clean_data:
        index = cell.column - 1
        header = str(clean_header[index].value).lower()
        value = str(cell.value) if cell.value else None
        if header in keywords:
            row_dict[header] = value
        elif re.match(r'^[12][0-9]{3}[13]$', header):
            if value: row_dict['periodos'].append((header, value[0:2]))
        else:
            raise Exception(f'Campo "{header}" no es reconocido')
    # ordenar periodos por fecha
    row_dict['periodos'].sort(key=lambda x: x[0])
    return row_dict

def clean_row(row):
    clean = []
    for cell in row:
        if cell.value is not None:
            clean.append(cell)
    return clean